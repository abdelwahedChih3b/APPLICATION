from weakref import ref
import openpyxl 
import os


def getPds(rownum,sheetname):
    pdsRow = []
    for i in range(4,105):
        pdsRow.append(sheetname.cell(row=rownum, column=i).value)
    return pdsRow

def sum2lists(list1,list2):

    zipped_lists = zip(list1, list2)
    sum = [x + y for (x, y) in zipped_lists]
    return sum





wb = openpyxl.load_workbook("pds001.xlsx")
niv0 = wb["n0"]
niv1 = wb["niv1"]
niv2 = wb["niv2"]
niv3 = wb["niv3"]
niv4 = wb["niv4"]
niv_1 = wb["niv-1"]


lastRowNiv0 = len(niv0["A"])
lastRowNiv1 = len(niv1["A"])
lastRowNiv2 = len(niv2["A"])
lastRowNiv3 = len(niv3["A"])
lastRowNiv4 = len(niv4["A"])
lastRowNiv_1 = len(niv_1["A"])
header = [niv0.cell(row=1,column=i).value for i in range(4,105)]
res = openpyxl.Workbook()
result = res.create_sheet(title="resultat finale")
result.append(header)

header = [niv0.cell(row=1,column=i).value for i in range(4,105)]
#lvl 0 products (aka : produits finis)
for i in range(2,lastRowNiv_1+1):
    ref = niv_1.cell(row=i, column=1).value
    #ref = "F0425450"
    pds =getPds(i,niv_1)      #poste de charge du pf (assemblage)
    listref = []
    listref.append(ref)

   # niv0ref = []
   #composant premier niveau lvl 1
    for j in range(2,lastRowNiv0+1):
        if niv0.cell(row=j, column=1).value == ref:
            refPlusOne = niv0.cell(row=j , column=2).value      #new reference to look for
            #niv0ref.append(niv0.cell(row=j, column=1).value)
            listref.append(refPlusOne)
            pdsniv = getPds(j,niv0)
            pds = sum2lists(pds,pdsniv)
            #composant niveau 2

            #pb here

            for k in range(2,lastRowNiv1+1):
                if niv1.cell(row=k,column=1).value == refPlusOne : #somme one by one
                    pdsniv = getPds(k,niv1)
                    pds = sum2lists(pds,pdsniv)
                    refPlusTwo =  niv1.cell(row=k,column=2).value
                    listref.append(refPlusTwo)
                    #composant niveau 3
                    for l in range(2,lastRowNiv2+1):
                        if niv2.cell(row=l,column=1).value == refPlusTwo :
                            pdsniv = getPds(l,niv2)
                            pds = sum2lists(pds,pdsniv)
                            refPlusThree =  niv2.cell(row=l,column=2).value
                            listref.append(refPlusThree)
                            #composant niveau 4
                            for m in range(2,lastRowNiv3+1):
                                if niv3.cell(row=m,column=1).value == refPlusThree :
                                    pdsniv = getPds(m,niv3)
                                    pds = sum2lists(pds,pdsniv)
                                    refPlusFour =  niv3.cell(row=m,column=2).value
                                    listref.append(refPlusFour)
                                    for n in range(2,lastRowNiv4+1):
                                        if niv4.cell(row=n,column=1).value == refPlusFour :
                                            listref.append(niv4.cell(row=n,column=2).value)
                                            pdsniv = getPds(n,niv4)
                                            pds = sum2lists(pds,pdsniv)
                                            
    pds.insert(0,ref)
    result.append(pds)
    pds = []
   
res.save(filename="resultat.xlsx")
os.system(r"resultat.xlsx")

        




            
            





